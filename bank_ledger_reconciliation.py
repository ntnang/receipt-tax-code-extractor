import os
import pandas
import pathlib
from openpyxl import load_workbook
import yaml
import sys
import re
import numpy as np

# bas = bank account statement
# deb = day end balance
# rt = result template

def extract_bas_deb_by_time(directory: str) -> dict:
    for file_name in os.listdir(directory):
        if file_name.startswith("So phu Ngan hang") and (file_name.endswith("xls") or file_name.endswith("xlsx")):
            bas_file_path = os.path.join(directory, file_name)
            print(bas_file_path)
            bas = pandas.read_excel(bas_file_path, dtype=str)
            bas.insert(0, "transaction_date_time", pandas.to_datetime(bas.iloc[:, 1]))
            bas.insert(1, "transaction_date", bas["transaction_date_time"].dt.date)
            print(bas)
            deb_idx = bas.groupby("transaction_date")["transaction_date_time"].idxmax()
            print(bas.groupby("transaction_date")["transaction_date_time"].max())
            deb_per_date = bas.iloc[deb_idx, [0, 10]]
            print(deb_per_date)
    return dict(zip(deb_per_date.iloc[:, 0], deb_per_date.iloc[:, 1]))

def extract_bas_deb_by_order(directory: str, cfg) -> dict:
    for file_name in os.listdir(directory):
        if file_name.startswith("So phu Ngan hang") and (file_name.endswith("xls") or file_name.endswith("xlsx")):
            cfg_by_bank = get_configurations_by_bank(file_name, cfg)
            print(cfg_by_bank)

            if cfg_by_bank is None:
                return None

            bas_file_path = os.path.join(directory, file_name)
            print(bas_file_path)
            bas = pandas.read_excel(bas_file_path, dtype=str, usecols=cfg_by_bank["col-range"], skiprows=cfg_by_bank["skip-rows"], skipfooter=cfg_by_bank["skip-footers"])
            print(bas)
            bas.insert(0, "transaction_date", pandas.to_datetime(bas.iloc[:, cfg_by_bank["date-col-idx"]], format=cfg_by_bank["date-format"]).dt.date)
            deduplicated_transaction_dates =  bas["transaction_date"].drop_duplicates()
            deb_per_date = {}
            print(bas)
            
            if (deduplicated_transaction_dates.is_monotonic_decreasing):
                deb_per_date = bas.groupby("transaction_date", as_index=False).first().iloc[:, [0, cfg_by_bank["bal-col-idx"] + 1]]
            elif (deduplicated_transaction_dates.is_monotonic_increasing):
                deb_per_date = bas.groupby("transaction_date", as_index=False).last().iloc[:, [0, cfg_by_bank["bal-col-idx"] + 1]]
            else:
                return None
            
            print(deb_per_date)
    return dict(zip(deb_per_date.iloc[:, 0], deb_per_date.iloc[:, 1].str.replace(cfg_by_bank["thousand-separator"], "").astype("int64")))

def calculate_bas_deb(directory: str, cfg_by_bank: dict) -> dict:
    if cfg_by_bank is None:
        return None
    
    for file_name in os.listdir(directory):
        if file_name.startswith("So phu Ngan hang") and (file_name.endswith("xls") or file_name.endswith("xlsx")):
            cfg_by_bank = get_configurations_by_bank(file_name, cfg)
            print(cfg_by_bank)

            if cfg_by_bank is None:
                return None

            bas_file_path = os.path.join(directory, file_name)
            print(bas_file_path)

            wb = load_workbook(bas_file_path, data_only=True)
            ws = wb.active   # or wb["Sheet1"]

            original_balance = ws[cfg_by_bank["bal-cell-addr"]].value   # Excel-style reference
            if type(original_balance) is str:
                original_balance = extract_balance_in_text(original_balance)
            print(original_balance)

            bas = pandas.read_excel(bas_file_path, dtype=str, usecols=cfg_by_bank["col-range"], skiprows=cfg_by_bank["skip-rows"], skipfooter=cfg_by_bank["skip-footers"])
            print(bas)

            bas.insert(0, "transaction_date", pandas.to_datetime(bas.iloc[:, cfg_by_bank["date-col-idx"]], format=cfg_by_bank["date-format"]).dt.date)

            bas.iloc[:, cfg_by_bank["debit-col-idx"] + 1] = pandas.to_numeric(bas.iloc[:, cfg_by_bank["debit-col-idx"] + 1], errors="coerce").fillna(0).astype("int64")
            bas.iloc[:, cfg_by_bank["credit-col-idx"] + 1] = pandas.to_numeric(bas.iloc[:, cfg_by_bank["credit-col-idx"] + 1], errors="coerce").fillna(0).astype("int64")

            # For debugging purpose
            # print(bas[(bas["transaction_date"] == pandas.to_datetime("2025-02-04").date()) & (bas.iloc[:, cfg_by_bank["debit-col-idx"] + 1] != 0)])
            # print(bas.iloc[:, cfg_by_bank["debit-col-idx"] + 1].dtype)

            total_debit_amount_per_day = bas.groupby("transaction_date", as_index=False)[bas.columns[cfg_by_bank["debit-col-idx"] + 1]].sum()
            total_credit_amount_per_day = bas.groupby("transaction_date", as_index=False)[bas.columns[cfg_by_bank["credit-col-idx"] + 1]].sum()

            total_debit_amount_per_day["cummulative_sum"] = total_debit_amount_per_day.iloc[:, 1].cumsum()
            total_credit_amount_per_day["cummulative_sum"] = total_credit_amount_per_day.iloc[:, 1].cumsum()

            print(total_debit_amount_per_day)
            print(total_credit_amount_per_day)

            deb_per_date = dict(zip(total_credit_amount_per_day["transaction_date"], original_balance + total_credit_amount_per_day["cummulative_sum"] - total_debit_amount_per_day["cummulative_sum"]))
            print(deb_per_date)

            return deb_per_date
    return None

def extract_balance_in_text(text: str) -> dict:
    # Extract digits, commas, and decimal point
    match = re.search(r'[\d,]+(?:\.\d+)?', text)
    if match:
        number_str = match.group(0).replace(',', '')
        return np.int64(float(number_str))
    return None

def extract_evn_deb(directory: str) -> dict:
    for file_name in os.listdir(directory):
        if file_name.startswith("So TGNH") and (file_name.endswith("xls") or file_name.endswith("xlsx")):
            evn_file_path = os.path.join(directory, file_name)
            print(evn_file_path)
            evn = pandas.read_excel(evn_file_path, header=None, dtype=str, skiprows=17, skipfooter=8)
            deb_per_date = evn.loc[evn[7].notna()].iloc[:, [4, 7]]
            deb_per_date[4] = pandas.to_datetime(deb_per_date[4].str[-10:], format="%d/%m/%Y").dt.date
            print(deb_per_date)
    return dict(zip(deb_per_date.iloc[:, 0], deb_per_date.iloc[:, 1].str.replace(" ", "").astype("int64")))

def export_results(bas_deb: dict, evn_deb: dict):
    if bas_deb is None or evn_deb is None:
        return

    rt_file_name = "KQ doi soat So phu NH - EVN_CM_009.xlsx"

    rt_wb = load_workbook(rt_file_name)
    rt_ws = rt_wb.active

    # print(bas_deb)
    # print(evn_deb)
    
    results = []
    for idx, (key, value) in enumerate(bas_deb.items()):
        matching_deb = evn_deb[key] if key in evn_deb else 0
        results.append(dict(index=idx+1, transaction_date=key, bas_deb_res=value, evn_deb_res=matching_deb, diff=(abs(matching_deb - value))))
    print(results)

    start_row = 8
    for i, row in enumerate(results, start=start_row):
        for j, key in enumerate(row, start=1):
            rt_ws.cell(row=i, column=j, value=row[key])
    
    rt_wb.save("KQ doi soat So phu NH - EVN_CM_009_final_result.xlsx")

    return results

def get_configurations_by_bank(file_name: str, cfg: dict) -> dict:
    normalized_file_name = re.sub(r"\s+", "", file_name).lower()
    for (key, value) in cfg.items():
        if key in normalized_file_name:
            return value
    return None

def get_configurations(file_name: str):
    # Load YAML file
    with open(file_name, 'r') as config:
        return yaml.safe_load(config)

cfg = get_configurations("configurations.yaml")
print(cfg)

bas_deb = calculate_bas_deb(pathlib.Path(__file__).parent.resolve(), cfg)
evn_deb = extract_evn_deb(pathlib.Path(__file__).parent.resolve())
export_results(bas_deb, evn_deb)

# exe_path = sys.argv[0]
# exe_dir = os.path.dirname(exe_path)
# bas_deb = extract_bas_deb_by_order(exe_dir, cfg)
# evn_deb = extract_evn_deb(exe_dir)
# export_results(bas_deb, evn_deb)

# bas_deb = extract_bas_deb_by_order(pathlib.Path(__file__).parent.resolve(), cfg)
# evn_deb = extract_evn_deb(pathlib.Path(__file__).parent.resolve())
# export_results(bas_deb, evn_deb)