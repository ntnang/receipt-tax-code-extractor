import PyPDF2
#import pathlib
import sys
import os
import re
import openpyxl
from openpyxl.styles import PatternFill
import pandas
from datetime import datetime

def extract_tax_codes_from_pdf(pdf_path, start_page = 0):
    tax_codes = []
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)

        # Get the number of pages in the PDF
        num_pages = len(pdf_reader.pages)

        # Ensure the start_page is within the valid range
        start_page = max(min(start_page, num_pages), 0)

        # Iterate through each page
        for page_number in range(start_page, num_pages):
            # Get the page
            page = pdf_reader.pages[page_number]

            # Extract text from the page
            text = page.extract_text()

            matches = re.findall(r'(?<![a-zA-Z0-9])\d{10}(?:-\d{3})?(?![a-zA-Z0-9])', text)

            tax_codes.extend([match for match in matches if (match != '0300942001-022')])
            print(f"{tax_codes}")

            # Print the extracted text for the current page
            #print(f"Page {page_number + 1} text:\n{text}\n")
        pdf_file.close()
    return tax_codes
            

def get_pdf_files(directory):
    pdf_files = []

    # Iterate through all files in the directory
    for file_name in os.listdir(directory):
        # Check if the file has a ".pdf" extension
        if file_name.endswith(".pdf"):
            # Build the full path to the PDF file
            pdf_path = os.path.join(directory, file_name)
            
            # Add the PDF file path to the list
            pdf_files.append(pdf_path)

    return pdf_files

def get_blacklist_tax_codes(directory):
    blacklist_tax_codes = []
    for file_name in os.listdir(directory):
        if file_name == "blacklist_tax_codes.xls" or file_name == "blacklist_tax_codes.xlsx":
            
            # Load the Excel file
            blacklist_path = os.path.join(directory, file_name)
            dataframe = pandas.read_excel(blacklist_path, sheet_name=None) # `sheet_name=None` loads all sheets into a dict

            # Iterate through all sheets
            for sheet_name, sheet_data in dataframe.items():
                print(f"Sheet: {sheet_name}")
                # Loop through rows of the DataFrame
                for index, row in sheet_data.iterrows():
                    if not row.isnull().all():
                        print(f"Row {index}:")
                        for col_name, cell_value in row.items():
                            if pandas.notnull(cell_value):
                                print(f"  Column '{col_name}': {cell_value}")
                                blacklist_tax_codes.append(cell_value)
    return blacklist_tax_codes


def export_to_excel(extracted_tax_codes, blacklist_tax_codes):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the active sheet (default is the first sheet)
    sheet = workbook.active

    sheet["A1"] = "Tax code"
    sheet["B1"] = "File name"

    # Starting row to write the data
    start_row = 2

    # Define a color fill (e.g., yellow background)
    yellow_background = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Loop through the list of lists and write it to consecutive rows and columns
    for row_index, tax_code in enumerate(extracted_tax_codes):
        value_cell = sheet.cell(row=start_row + row_index, column=1, value=tax_code["value"])
        sheet.cell(row=start_row + row_index, column=2, value=tax_code["file_name"])
        if tax_code["value"] in blacklist_tax_codes:
            value_cell.fill = yellow_background

    try:
        # Save the workbook to a file
        workbook.save('checked_extracted_tax_codes.xlsx')
    except PermissionError:
        print("Permission error: The file might be open or locked.")
    except Exception as e:
        print(f"An error occurred: {e}")

    print('Data written to the Excel file within a loop successfully.')

# Path to the current script
#current_path = pathlib.Path(__file__).parent.resolve()

# Get the path to the executable
exe_path = sys.argv[0]

# Get the directory containing the executable
exe_dir = os.path.dirname(exe_path)

blacklist_tax_codes = get_blacklist_tax_codes(exe_dir)

# File is automatically closed when you exit the 'with' block

pdf_files = get_pdf_files(exe_dir)
extracted_tax_codes_with_relevent_file_names = []

# Extracting PDF files
for pdf_file in pdf_files:
    print(f"Extracting tax codes in: {pdf_file}")
    extracted_tax_codes = extract_tax_codes_from_pdf(pdf_file)
    for tax_code in extracted_tax_codes:
        pdf_file_name = os.path.basename(pdf_file)
        extracted_tax_codes_with_relevent_file_names.append({ "value": tax_code, "file_name": pdf_file_name })

print("----------blacklist_tax_codes------------")
print(blacklist_tax_codes)
print("----------extracted_tax_codes_with_relevent_file_names------------")
print(extracted_tax_codes_with_relevent_file_names)

export_to_excel(extracted_tax_codes_with_relevent_file_names, blacklist_tax_codes)

#Open a file in append mode ('a')
with open('log.txt', 'a') as file:
    logs = []

    logs.append("--------------------------------------------------------")

    # Get the current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    logs.append(current_datetime)

    logs.append("----------exe_dir------------")
    logs.append(exe_dir)

    logs.append("----------pdf_files------------")
    logs += pdf_files

    logs.append("----------blacklist_tax_codes------------")
    logs += blacklist_tax_codes

    logs.append("----------extracted_tax_codes_with_relevent_file_names------------")
    logs.append(str(extracted_tax_codes_with_relevent_file_names))

    logs.append("--------------------------------------------------------")

    # Add a newline character at the end of each line
    logs = [log + "\n" for log in logs]

    print(logs)

    # Write content to the file
    file.writelines(logs)

input("Press Enter to continue.")
